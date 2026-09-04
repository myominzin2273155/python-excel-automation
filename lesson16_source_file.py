import os
import openpyxl

source_folder = "sales_reports"
target_folder = "merged_results"

os.makedirs(target_folder, exist_ok=True)


master_wb = openpyxl.Workbook()
master_sheet = master_wb.active
master_sheet.title = "Consolidated Data"

headers = ["Source File", "Sales Rep", "Region", "Sales Amount", "Bonus"]
master_sheet.append(headers)

if os.path.exists(source_folder):
    for file_name in os.listdir(source_folder):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(source_folder, file_name)

            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = [file_name] + list(row)
                master_sheet.append(row_data)

    output_file = os.path.join(target_folder, "master_with_source.xlsx")
    master_wb.save(output_file)
    print("SUCCESS: Consolidated all files with Source File names!")
else:
    print(f"ERROR: '{source_folder} folder no")    
    




