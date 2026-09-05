import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

target_folder = "merged_results"
file_path = os.path.join(target_folder, "master_sorted_filtered.xlsx")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)

    if "Top Performers" in wb.sheetnames:
        sheet = wb["Top Performers"]

        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"

        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical = "center")

        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3"),
        )
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border

                if cell.column == 4:
                    cell.number_format = "#,##0"
                    cell.alignment = right_align
                else:
                    cell.alignment = center_align

        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 5, 12)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Top Performers Sales Comparison"
        chart.y_axis.title = "Sales Amount (MMK)"
        chart.x_axis.title = "Sales Rep"

        data = Reference(
            sheet, min_col=4, min_row=1, max_row=sheet.max_row
        )

        categories = Reference(
            sheet, min_col=2, min_row=2, max_row=sheet.max_row
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, "F2")

        wb.save(file_path)
        print(
            "SUCCESS: Excel formatting and Bar Chart generated successfully!"
        )
    else:
        print("ERROR. 'Top Performers' sheet not found in the workbook!")
else:
    print(
        "ERROR: File not found! Please run lesson17_sorting_filtering.py first."
    )                                    