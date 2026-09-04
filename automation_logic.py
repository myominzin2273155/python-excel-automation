import openpyxl
import os

base_dir = os.path.dirname(os.path.abspath(__file__))
reports_folder = os.path.join(base_dir, "all_reports")
output_file = os.path.join(reports_folder, "merged_master_report.xlsx")

merged_wb = openpyxl.Workbook()
merged_ws = merged_wb.active
merged_ws.title = "Master_Data"

header_saved = False
header_row = []
all_data_rows = []

for filename in os.listdir(reports_folder):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and filename != "merged_master_report.xlsx":
        file_path = os.path.join(reports_folder, filename)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        if not header_saved:
            header = [cell.value for cell in ws[1]]
            merged_ws.append(header)
            header_saved = True

        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                all_data_rows.append(list(row))

all_data_rows.sort(key=lambda x: int (x[2]) if x[2] is not None else 0, reverse=True)
merged_ws.append(header_row)

for data in all_data_rows:
    merged_ws.append(data)

merged_wb.save(output_file)

print("SUCCESS: Amount sorts ")