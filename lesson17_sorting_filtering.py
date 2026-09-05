import os
import openpyxl

source_folder = "sales_reports"
target_folder = "merged_results"

os.makedirs(target_folder, exist_ok=True)

master_wb = openpyxl.Workbook()

all_sheet = master_wb.active
all_sheet.title = "All Sales Data"

top_sheet = master_wb.create_sheet(title="Top Performers")

headers = ["Source File", "Sales Rep", "Department", "Sales Amount"]
all_sheet.append(headers)
top_sheet.append(headers)

filtered_rows =[]

if os.path.exists(source_folder):
    for file_name in os.listdir(source_folder):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(source_folder, file_name)
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                row_data = [file_name] + list (row)
                all_sheet.append(row_data)

                if len(row_data) > 3 and row_data[3] is not None:
                    try:

                        val_str = str(row_data[3]).replace(",", "").strip()
                        sales_val = float(val_str)

                        if sales_val >= 3000000:
                            filtered_rows.append((sales_val, row_data))
                    except ValueError:
                        pass
filtered_rows.sort(key=lambda x: x[0], reverse=True)

for sales_val, row_data in filtered_rows:
    top_sheet.append(row_data)

output_file = os.path.join(target_folder, "master_sorted_filtered.xlsx")
master_wb.save(output_file)

print(f"SUCCESS: {len(filtered_rows)} rows added to Top Performers Sheet!")