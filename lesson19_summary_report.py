import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "merged_results", "master_sorted_filtered.xlsx")

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)

    if "Summary Report" in wb.sheetnames:
        del wb["Summary Report"]

        Summary_sheet = wb.create_sheet(title="Summary Report", index=0)

        report_data = []

        for sheet_name in wb.sheetnames:
            if sheet_name == "Summary Report":
                continue

            sheet = wb[sheet_name]
            total_sales = 0
            total_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[3] is not None:
                    try:
                        total_sales += float(row[3])
                        total_count += 1
                    except ValueError:
                        pass
            report_data.append([sheet_name, total_count, total_sales])

            Summary_sheet.append(["Sheet Name", "Total Transactions", "Total Sales Amount"])

            for row in report_data:
                Summary_sheet.append(row)

            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=11)

            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"),
                right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"),
                bottom=Side(style="thin", color="D3D3D3"),
            )

            for cell in Summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in Summary_sheet.iter_rows(min_row=2, max_row=Summary_sheet.max_row):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border

                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == 2:
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == 3:
                        cell.number_format = "#,##0"
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                for col in Summary_sheet.columns:
                    max_len = max(len(str(cell.value or "")) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    Summary_sheet.column_dimensions[col_letter].width = max(max_len + 6, 18)

                wb.save(file_path)
                print("SUCCESS: Multi-sheet Summary Report generated succellfully!")

            else:
                print(
                    "ERROR: File not found! Please make sure master_sorted_filtered.xlsx exists."
                )                                            